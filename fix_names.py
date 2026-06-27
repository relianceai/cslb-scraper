#!/usr/bin/env python3
"""Split the scraped `last_name` column into `middle name` + `last name`.

The CSLB scraper put the middle name (when present) and the last name together in
one `last_name` column, e.g. "WAYNE KEENE". This script splits each value so the
final word becomes the real `last name` and everything before it becomes the
`middle name` (empty when there is only one word).

Reads each input .xlsx, writes a `fixed_<name>.xlsx` copy, preserving every other
column. Does not modify the originals.
"""
import sys
import openpyxl


SUFFIXES = {"JR", "SR", "II", "III", "IV", "V"}


def split_name(value):
    """Return (middle, last, suffix) from a combined last-name string.

    The final word is the surname, unless it is a recognised generational suffix
    (JR/SR/II/III/IV/V) in which case it is split out and the preceding word is the
    surname. Everything before the surname becomes the middle name.
    """
    if value is None:
        return None, None, None
    text = str(value).strip()
    if not text:
        return None, None, None
    parts = text.split()           # collapses runs of whitespace

    suffix = None
    if len(parts) > 1 and parts[-1].upper().strip(".") in SUFFIXES:
        suffix = parts.pop()       # keep original token form (e.g. "JR.")

    last = parts[-1]
    middle = " ".join(parts[:-1])  # "" when no middle name
    return (middle or None), last, suffix


def fix_file(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # locate the last_name column by header
    header = [c.value for c in ws[1]]
    try:
        ln_idx = header.index("last_name") + 1  # 1-based
    except ValueError:
        raise SystemExit(f"{path}: no 'last_name' column found (headers: {header})")

    # rename existing column -> "middle name"; two new columns after it
    last_idx = ln_idx + 1
    suffix_idx = ln_idx + 2
    ws.insert_cols(last_idx, amount=2)
    ws.cell(row=1, column=ln_idx).value = "middle name"
    ws.cell(row=1, column=last_idx).value = "last name"
    ws.cell(row=1, column=suffix_idx).value = "suffix"

    for r in range(2, ws.max_row + 1):
        original = ws.cell(row=r, column=ln_idx).value
        middle, last, suffix = split_name(original)
        ws.cell(row=r, column=ln_idx).value = middle
        ws.cell(row=r, column=last_idx).value = last
        ws.cell(row=r, column=suffix_idx).value = suffix

    out = path.rsplit("/", 1)
    out[-1] = "fixed_" + out[-1]
    out_path = "/".join(out)
    wb.save(out_path)
    return out_path, ws.max_row - 1


if __name__ == "__main__":
    for p in sys.argv[1:]:
        out_path, n = fix_file(p)
        print(f"{p} -> {out_path}  ({n} rows)")
